import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
workbook = load_workbook(r"C:\Users\pragy\OneDrive\Desktop\dsa\pylint.xlsx")

# Select the worksheet to work on
worksheet = workbook['Sheet1']
df = pd.DataFrame(columns=['Module','Score', 'Path', 'Finding'])
for i in df.index:
    print(df['Score'][i])



# Merge cells A1 to C1 and center the content
#worksheet.merge_cells('A2:A8')
#worksheet['A2'].alignment = Alignment(horizontal='center', vertical='center')

# Save the changes to the workbook
workbook.save('example.xlsx')