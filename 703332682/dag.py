''' Converting the result from pylint in an excel format'''
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
lis = []
# create an empty DataFrame
df = pd.DataFrame(columns=['Module', 'Score', 'Path', 'Finding', 'count','Diff'])

# open the file and read the lines
with open('./myfile.txt', 'r', encoding='utf-16') as f:
    lines = f.readlines()

# loop through the lines and extract the information
for line in lines:
    result = re.match(r'Your', line)
    if bool(result):
        df.loc[len(df)] = ["", line.strip(), "", "", "",""]
    else:
        module_match = re.match(r'\*{13}\sModule\s(\S*)', line)
        if module_match:
            module = module_match.group(1)+'.py'
            lis.append(module)
        idx = line.find(":")
        if idx != -1:
            finding = line[idx+1:].strip()
            # print(repo)
        # append the row to the DataFrame
            if len(lis) == 1:
                df.loc[len(df)] = [lis[0], "", "", finding, "",""]
                lis.pop()
            else:
                df.loc[len(df)] = ["", "", "", finding, "",""]
df.to_excel('hail.xlsx', index=False)

workbook = load_workbook(r"C:\Users\pragy\OneDrive\Desktop\dsa\hail.xlsx")

# Select the worksheet to work on
curr, t = 2, 2
Z = 0
worksheet = workbook['Sheet1']
for i in df.index:
    SCORE = str(df['Score'][i])
    if len(SCORE) != 0:
        res = re.match(
            r'Your\scode\shas\sbeen\srated\sat\s(\d+\.\d+)', str(df['Score'][i]))
        # print(res.group(1))
        if res and res.group(1) == '10.00':
            Z += 1
            worksheet[f'B{i+2}'].value = ""
            worksheet[f'E{i+2}'].value = None
            continue
        # Merge cells and center the content
        # print(curr,i+1)
        worksheet[f'B{curr+Z}'].value = worksheet[f'B{i+2}'].value
        worksheet[f'B{i+2}'].value = ""
        print(worksheet[f'A{curr+Z}'].value)
        worksheet.merge_cells(f'A{curr+Z}:A{i+1}')
        worksheet[f'A{curr+Z}'].alignment = Alignment(
            horizontal='center', vertical='center')
        worksheet.merge_cells(f'B{curr+Z}:B{i+1}')
        worksheet[f'B{curr+Z}'].alignment = Alignment(
            horizontal='center', vertical='center')
        worksheet[f'E{curr+Z}'].value = (i+2-(curr+Z))
        worksheet[f'E{i+2}'].value = None
        worksheet.merge_cells(f'E{curr+Z}:E{i+1}')
        worksheet[f'E{curr+Z}'].alignment = Alignment(
            horizontal='center', vertical='center')
        curr = i+3
        Z = 0

worksheet['C2'].value = "LD-CORA-CLL-API"
worksheet.merge_cells(f'C{t}:C{curr-1}')
worksheet['C2'].alignment = Alignment(horizontal='center', vertical='center')
# print(curr)


# for row in dataframe_to_rows(df, index=False, header=True):
#   worksheet.append(row)

# auto-fit columns to fit the contents of the cells
for col in worksheet.columns:
    MAX_LENGTH = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > MAX_LENGTH:
                MAX_LENGTH = len(str(cell.value))
                # print(MAX_LENGTH,cell.value)
        except ValueError:
            pass
    ADJUSTED_WIDTH = MAX_LENGTH + 2
    worksheet.column_dimensions[column].width = ADJUSTED_WIDTH
workbook.save('hail.xlsx')


# Get the values from the 'count' column
count_values = [cell.value for cell in worksheet['E'] if cell.value is not None]
#print(count_values)

#workbook1 = load_workbook(r"C:\Users\pragy\OneDrive\Desktop\dsa\pylint.xlsx")


#worksheet1 = workbook1['Sheet1']
#count_values1 = [str(cell.value+'33' for cell in worksheet1['E'] if cell.value is not None]
#print(count_values1)

cell_number = data[data[column_name] == search_value].index[0] + 2

